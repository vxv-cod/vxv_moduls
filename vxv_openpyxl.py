'''Работа с файлами XLSX при помощи модуля openpyxl'''
'''https://docs-python.ru/packages/modul-openpyxl/'''

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# создаем книгу 
wb = Workbook()
# делаем единственный лист активным 
ws = wb.active
# Адресс файла загрузки
path = r'C:\mydata\generator\py\zszxz.xlsx'
# Загрузить книгу
wb = openpyxl.load_workbook(path)



# Вставить рабочий лист в конец (по умолчанию)
sheet = wb.create_sheet("Mysheet")
# Вставить рабочий лист в первую позицию
sheet = wb.create_sheet("Mysheet", 0)
# Вставить рабочий лист в предпоследнюю позицию
ws3 = wb.create_sheet("Mysheet", -1)
# Переименовать название листа
sheet.title = "NewPage"
# Цвет фона вкладки с этим заголовком по умолчанию белый.
# Можно изменить этот цвет, указав цветовой код RRGGBB для атрибута листа Worksheet.sheet_properties.tabColor:
sheet.sheet_properties.tabColor = "1072BA"
# Рабочий лист можно получить, используя его имя в качестве ключа 
# экземпляра созданной книги Excel:
sheet = wb['sheet_name']
# Копия листа
sheet = ws['zszxz']
cp_sheet = wb.copy_worksheet(sheet)
# Получаем названия всех вкладок
print(wb.sheetnames)
# Удалить лист
wb.remove(sheet)

# Максимальное количество заполненных строк
max_row = ws.max_row
# Максимальное количество заполненных строк
max_column = ws.max_column

# Получить значение ячейки
cell = sheet['A1']
val = cell.value
val = sheet['A1'].value
# Получить номер столбца по ячейке
num_coll = ws['C4'].column
# Получить номер строки по ячейке
num_row = ws['C4'].row

# Переназначить
cell = sheet['A1']
cell.value = 'zszxz'
sheet['A1'].value = 'zszxz'
# или написать текстовую форму
cell = 'zszxz'
sheet['A1'] = 'zszxz'
# или через индексы ячейки
# получить
cell = sheet.cell(row=1, column=1)
cell.value = 'zszxz'
sheet.cell(row=1, column=1).value = 'zszxz'
# назначить без потери форматирование ячейки в шаблоне
cell = sheet.cell(row=1, column=1, value='zszxz')
cell = sheet.cell(4, 1, 'dd')
# Написать формулу
sheet['A3'] = '=SUM(1, 1)'


# Получите одну ячейку:
cell = sheet['A1']
cell = sheet.cell(row=1, column=1)
sheet.cell(row=1, column=1, value=None)


# Получить указанный диапазон строк
cells_range = sheet['A1':'C1']
for cells in cells_range:
    for cell in cells:
        print(cell)

# Аналогично можно получить диапазоны имеющихся строк или столбцов на листе:
# Все доступные ячейки в колонке `C`
colC = ws['C']
# Все доступные ячейки в диапазоне колонок `C:D`
col_range = ws['C:D']
# Все доступные ячейки в строке 10
row10 = ws[10]
# Все доступные ячейки в диапазоне строк `5:10`
row_range = ws[5:10]

# Можно также использовать метод Worksheet.iter_rows():
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
   for cell in row:
       print(cell)
# Таким же образом sheet.columns получает все столбцы и не повторяет их:
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in col:
        print(cell)
'''Примечание. Из соображений производительности метод Worksheet.iter_cols() 
недоступен в режиме только для чтения.'''

'''Режим только для чтения'''
'''# Важно. Рабочая книга, в таком режиме должна быть явно закрыта с помощью метода Workbook.close()'''
from openpyxl import load_workbook
# открываем файл только для чтения
wb = load_workbook(filename='big_data.xlsx', read_only=True)
# открываем лист по его имени
ws = wb['Sheet']
for row in ws.rows:
    for cell in row:
        print(cell.value)
# закрываем книгу после прочтения
wb.close()

'''Режим только для записи'''
from openpyxl import Workbook
# создаем книгу только для записи
wb = Workbook(write_only=True)
ws = wb.create_sheet()
# заполним 1000 строк х 100 столбцов
for i in range(1, 1000):
    ws.append([f'{i*j}' for j in range(1, 100)])
# сохраним файл
wb.save('big_data.xlsx')




'''Если необходимо перебрать все строки или столбцы файла,
то можно использовать свойство Worksheet.rows:'''
'''Получить все строки'''
rows = ws.rows
for row in ws.rows:
    print(row)
'''или'''
rows = tuple(ws.rows)
# Таким же образом sheet.columns получает все столбцы и не повторяет их
'''# Получить все столбцы'''
cols = sheet.columns
'''или'''
columns = tuple(ws.columns)

'''Получение только значений ячеек активного листа.
Если просто нужны значения из рабочего листа, то можно использовать свойство 
активного листа Worksheet.values. Это свойство перебирает все строки на листе, 
но возвращает только значения ячеек:'''
for row in ws.values:
   for value in row:
     print(value)
# или
values = tuple(ws.values)
print(values)

r'''Добавление данных в ячейки листа списком.
Модуль openpyxl дает возможность супер просто и удобно добавлять данные в конец листа электронной таблицы. Такое удобство обеспечивается методом объекта листа Worksheet.append(iterable), где аргумент iterable - это любой итерируемый объект (список, кортеж и т.д.). Такое поведение позволяет, без костылей, переносить в электронную таблицу данные из других источников, например CSV файлы, таблицы баз данных, дата-фреймы из Pandas и т.д.
Метод Worksheet.append() добавляет группу значений в последнюю строку, которая не содержит данных.
Если это список: все значения добавляются по порядку, начиная с первого столбца.
Если это словарь: значения присваиваются столбцам, обозначенным ключами (цифрами или буквами).
Варианты использования:
добавление списка: .append(["ячейка A1", "ячейка B1", "ячейка C1"])
добавление словаря:
вариант 1: .append({"A" : "ячейка A1", "C" : "ячейка C1"}), в качестве ключей используются буквы столбцов.
вариант 2: .append({1 : "ячейка A1", 3 : "ячейка C1"}), в качестве ключей используются цифры столбцов.
'''
data = [
    ["ячейка A1", "ячейка B1", "ячейка C1"],
    ["ячейка A2", "ячейка B2", "ячейка C2"]
]
for row in data:
    ws.append(row)


'''Сохранение данных книги в виде потока.
Если необходимо сохранить файл в поток, например, при использовании веб-приложения, 
такого как Flask или Django, то можно просто предоставить tempfile.NamedTemporaryFile():'''
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
wb = Workbook()
with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()

'''# Сохранить
# Внимание. Эта операция перезапишет существующий файл без предупреждения!!!'''
wb.save(path)

'''Можно указать атрибут template=True, чтобы сохранить книгу как шаблон:'''
from openpyxl import load_workbook
wb = load_workbook('test.xlsx')
wb.template = True
wb.save('test_template.xltx')

'''Примечание. Атрибут wb.template по умолчанию имеет значение False, 
это означает - сохранить как документ.
Внимание. Следующее не удастся:'''
from openpyxl import load_workbook
wb = load_workbook('test.xlsx')
# Необходимо сохранить с расширением *.xlsx
wb.save('new_test.xlsm') # MS Excel не может открыть документ

# Нужно указать атрибут `keep_vba=True`
wb = load_workbook('test.xlsm')
wb.save('new_test.xlsm')

wb = load_workbook('test.xltm', keep_vba=True)
# Если нужен шаблон документа, то необходимо указать расширение *.xltm.
wb.save('new_test.xlsm') # MS Excel не может открыть документ

'''Есть несколько флагов, которые можно использовать в функции openpyxl.load_workbook().
data_only: определяет, будут ли содержать ячейки с формулами - формулу (по умолчанию) 
    или только значение, сохраненное/посчитанное при последнем чтении листа Excel.
keep_vba определяет, сохраняются ли какие-либо элементы Visual Basic (по умолчанию). 
    Если они сохранены, то они не могут изменяться/редактироваться.'''


'''-----------------------------------------------------------------------------------------'''
'''Работа с ячейками'''
'''-----------------------------------------------------------------------------------------'''

def importdata(min_row):
    # cells = [row for row in ws.iter_rows(min_row=min_row, max_col=ws.max_column, max_row=ws.max_row)]
    # cells = tuple(ws.iter_rows(min_row=min_row, max_col=ws.max_column, max_row=ws.max_row))

    cells = tuple(ws.rows)
    # cells = tuple(ws.columns)[0]

    vals = [[cell.value for cell in cells[row]] for row in range(len(cells))]
    return vals


'''Работа со стилями текста и ячеек, модуль openpyxl в Python'''
'''https://docs-python.ru/packages/modul-openpyxl/rabota-stiljami/'''

'''Шифр корпоративного желтого цвета - #FFD200'''
'''Подобрать нужный код цвета можено здедсь: https://color2.ru/000000'''

'''Ниже приведены значения по умолчанию установленные модулем openpyxl:'''
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, GradientFill

# СТИЛЬ ШРИФТА
font = Font(
        name='Calibri',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
            )

# ЗАЛИВКА ЯЧЕЕК
fill = PatternFill(fill_type=None, fgColor='FFFFFFFF')

# ГРАНИЦЫ ЯЧЕЕК
border = Border(
            left=Side(border_style=None, color='FF000000'),
            right=Side(border_style=None, color='FF000000'),
            top=Side(border_style=None, color='FF000000'),
            bottom=Side(border_style=None, color='FF000000'),
            diagonal=Side(border_style=None, color='FF000000'),
            diagonal_direction=0, 
            outline=Side(border_style=None, color='FF000000'),
            vertical=Side(border_style=None, color='FF000000'),
            horizontal=Side(border_style=None, color='FF000000')
               )

# ВЫРАВНИВАНИЕ В ЯЧЕЙКАХ
alignment=Alignment(
                horizontal='general',
                vertical='bottom',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0
                   )

horizontal_alignments = (
    "general", "left", "center", "right", "fill", "justify", "centerContinuous",
    "distributed", )
vertical_aligments = (
    "top", "center", "bottom", "justify", "distributed",
)

'''Cтили ячеек электронной таблицы.
Существует два типа стилей: стили ячеек и именованные стили, также известные как шаблоны стилей.
Стили ячеек являются общими для объектов, и после того, как они были назначены, их нельзя изменить. 
Это предотвращает нежелательные побочные эффекты, такие как изменение стиля для большого количества 
ячеек при изменении только одной.'''
# Например:

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'].value = 'Ячейка `A1`'
ws['D4'].value = 'Ячейка `D4`'
# задаем стиль шрифта текста - цвет  ячейке
ft = Font(color="FF0000")
# применяем стиль к ячейкам
ws['A1'].font = ft
ws['D4'].font = ft
# это изменение не сработает
ws['D4'].font.italic = True 
# Если необходимо изменить шрифт, 
# то его необходимо переназначить новым стилем
ws['A1'].font = Font(color="FF0000", italic=True)
wb.save('test.xlsx')

'''Создания нового стиля на основе другого.
Модуль openpyxl поддерживает копирование стилей.
Пример создания нового стиля на основе другого:'''
from openpyxl.styles import Font
from copy import copy
# задаем стиль
ft1 = Font(name='Arial', size=14)
# копируем стиль
ft2 = copy(ft1)
# а вот теперь на основе скопированного стиля 
# можно создать новый, изменив атрибуты 
ft2.name = "Tahoma"
# имя шрифта первого стиля
ft1.name
# 'Arial'
# имя шрифта нового стиля
ft2.name
# 'Tahoma'
# размер остался как у первого
ft2.size # copied from the
# 14.0


'''Применение стилей.
Стили применяются непосредственно к ячейкам.'''
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
wb = Workbook()
ws = wb.active
c = ws['A1']
c.value = 'Ячейка `A1`'
c.font = Font(size=12)
# можно напрямую
ws['A2'].value = 'Ячейка `A2`'
ws['A2'].font = Font(size=12, bold=True)
wb.save('test.xlsx')


'''Стили также могут применяться к столбцам и строкам, но обратите внимание, 
что это относится только к ячейкам, созданным (в Excel) после закрытия файла. 
Если необходимо применить стили ко всем строкам и столбцам, то нужно применить 
стиль к каждой ячейке самостоятельно.
Это ограничение формата файла:'''
col = ws.column_dimensions['A']
col.font = Font(bold=True)
row = ws.row_dimensions[1]
row.font = Font(underline="single")


'''Горизонтальное и вертикальное выравнивание текста.
Горизонтальное и вертикальное выравнивание в ячейках выставляется атрибутом ячейки .alignment и классом Alignment().
Пример горизонтального выравнивания текста:'''

# выравниваем текст в ячейках стилями
ws['A1'].alignment = Alignment(horizontal='left')
ws['A2'].alignment = Alignment(horizontal='center')
ws['A3'].alignment = Alignment(horizontal='right')

# Пример вертикального выравнивания данных в ячейке:
# объединим ячейки в диапазоне `B2:E2`
ws.merge_cells('B2:E2')
# в данном случае крайняя верхняя-левая ячейка это `B2`
megre_cell = ws['B2']
# запишем в нее текст
megre_cell.value = 'Объединенные ячейки `B2 : E2`'
# установить высоту строки
ws.row_dimensions[2].height = 30
# установить ширину столбца
ws.column_dimensions['B'].width = 40
# выравнивание текста
megre_cell.alignment = Alignment(horizontal="center", vertical="center")


'''Оформление границ ячеек.'''
'''Цвет и стиль границ/бордюров ячеек выставляется атрибутом ячейки .border и 
классом Border() совместно с классом Side().'''
'''При этом аргумент стиля границ ячеек border_style может принимать ОДИН из следующих значений: 
'dashDotDot', 'medium', 'dotted', 'slantDashDot', 'thin', 'hair', 'mediumDashDotDot', 'dashDot', 'double', 'mediumDashed', 'dashed', 'mediumDashDot' и 'thick'.
"Пунктирная точка", "средняя", "пунктирная точка", "наклонная точка", "тонкая", "волосы", "средняя точка", "пунктирная точка", "двойная", "средняя точка", "пунктирная точка", "средняя точка" и "толстая".

Пример стилизации границ одной ячейки:'''
from openpyxl import Workbook
from openpyxl.styles import Border, Side
wb = Workbook()
ws = wb.active
cell = ws['B2']
# определим стили сторон
thins = Side(border_style="medium", color="0000ff")
double = Side(border_style="dashDot", color="ff0000")
# рисуем границы
cell.border = Border(top=double, bottom=double, left=thins, right=thins)

thins = Side(border_style="thin", color="000000")
cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)


'''Заливка ячеек цветом и цвет текста.
Цвет заливки ячеек выставляется атрибутом ячейки .fill и классом PatternFill().'''

'''Обязательный аргумент fill_type (по умолчанию равен None) класса PatternFill() может принимать значения:
если fill_type='solid', то нужно обязательно указывать аргумент цвета заливки fgColor.
следующие значения аргумента fill_type применяются самостоятельно (без аргумента fgColor) и представляют собой предустановленные цвета заливки : 'darkHorizontal', 'lightDown', 'lightGray', 'darkDown', 'darkGrid', 'darkUp', 'darkGray', 'darkVertical', 'darkTrellis', 'mediumGray', 'lightVertical', 'lightTrellis', 'lightGrid', 'lightHorizontal', 'gray0625', 'lightUp', 'gray125'.
Внимание: если аргумент fill_type не указан, то fgColor не будет иметь никакого эффекта!
Пример заливки одной ячейки:'''
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
wb = Workbook()
ws = wb.active
# объединим ячейки в диапазоне `B2:E2`
ws.merge_cells('B2:E2')
megre_cell = ws['B2']
# запишем в нее текст
megre_cell.value = 'Объединенные ячейки `B2 : E2`'
# установить высоту строки
ws.row_dimensions[2].height = 30
# установить ширину столбца
ws.column_dimensions['B'].width = 40
# заливка ячейки цветом
megre_cell.fill = PatternFill('solid', fgColor="DDDDDD")
# шрифт и цвет текста ячейки 
megre_cell.font = Font(bold=True, color='FF0000', name='Arial', size=14)
# ну и для красоты выровним текст
megre_cell.alignment = Alignment(horizontal='center', vertical='center')
# сохраняем и смотрим что получилось
# wb.save("cell_color.xlsx")


'''Именованные стили NamedStyle.
В отличие от простых стилей ячеек, именованные стили изменяемы и используется для объединения в себе нескольких стилей, таких как шрифты, границы, выравнивание и т. д. Они имеют смысл, когда необходимо применить форматирование к множеству разных ячеек одновременно. Об именованных стилях можно думать как о классах CSS при оформлении HTML-разметки. Именованные стили регистрируются в рабочей книге.
Примечание. После назначения ячейке именованного стиля, дальнейшие/дополнительные изменения этого стиля не повлияют на стиль ячейки.
Как только именованный стиль зарегистрирован в рабочей книге, на него можно ссылаться просто по имени.
Создание именованного стиля.'''
from openpyxl.styles import NamedStyle, Font, Border, Side
# создание переменной именованного стиля
name_style = NamedStyle(name="highlight")
# применение стилей к созданной переменной
name_style.font = Font(bold=True, size=20)
bd = Side(style='thick', color="000000")
name_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# После создания именованного стиля его нужно зарегистрировать в рабочей книге:
wb.add_named_style(name_style)
# Именованные стили также будут автоматически зарегистрированы при первом назначении их ячейке:
ws['A1'].style = name_style
# После регистрации стиля в рабочей книге, применять его можно только по имени:
ws['D5'].style = 'highlight'


# Объединить ячейки 
sheet.merge_cells('A2:D2')
# Разъединить ячейки
sheet.unmerge_cells('A2:D2')

from openpyxl.drawing.image import Image
# Установить изображение
img = Image(r'C:\mydata\generator\py\main.jpg')
 # Установить описание ячейки изображения
sheet['A1'] = 'you are my angel'
 # Вставить картинку
sheet.add_image(img, 'A1')

# Скрытый ряд A-B
ws.column_dimensions.group('A', 'B', hidden=True)
# Скрытые столбцы 1-5 
ws.row_dimensions.group(1, 5, hidden=True)



'''Фиксация строк/столбцов в качестве шапки при помощи openpyxl'''
'''https://docs-python.ru/packages/modul-openpyxl/zakrepit-stroki-stolbtsy-shapku/'''
'''Закрепление строк/колонок в качестве шапки электронной таблицы.
Модуль openpyxl имеет возможность зафиксировать строки и колонки находящиеся выше и левее указанной ячейки, при помощи свойства листа электронной таблицы Worksheet.freeze_panes. Другими словами, это свойство фиксирует все что выше и левее указанной ячейки.
Например, если указать ws.freeze_panes = 'A3', то это заставит программу Excel зафиксировать только две верхние строки от прокрутки по вертикали. А если указать ws.freeze_panes = 'B3', то это зафиксирует две верхние строки и колонку 'A' от прокрутки по горизонтали.
Свойство Worksheet.freeze_panes должно быть вызвано после вставки некоторых данных. Смотрим пример фиксации трех первых строк листа, в качестве шапки, а также фиксации левого столбца от прокрутки по горизонтали:'''
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
#!!! СОЗДАЕМ ШАПКУ
# фиксируем все, что левее и выше ячейки "B4"
ws.freeze_panes = "B4"
# Для наглядности задаем и применяем стиль для `шапок`
ws.row_dimensions[3].font = Font(bold=True, name='Arial', size=10)
ws.column_dimensions['A'].font = Font(bold=True, name='Arial', size=10)
# сохраняем и смотрим результат
wb.save('freeze_panes.xlsx')


'''Печать шапки электронной таблицы на каждом листе.
Если электронная таблица очень большая и выходит за пределы печати одного листа (в том числе за правые границы), то для удобства сотрудника просматривающего документ, необходимо печатать шапку электронной таблицы на каждом листе документа.
Для печати "шапки таблицы" на каждом листе документа, необходимо указать диапазон строк, которые занимает шапка в свойстве листа Worksheet.print_title_rows. Например, выражение ws.print_title_rows = '2:3' заставить Excel печатать вверху каждого листа диапазон строк с 2 по 3 (т.е. 2-ю и 3-ю строки).
Для печати зафиксированных слева столбцов на каждом листе документа (! если электронная таблица выходит за пределы печати справа), необходимо в свойстве листа Worksheet.print_title_cols указать диапазон столбцов. Например, выражение ws.print_title_cols = 'A:B' заставить Excel печатать слева каждого листа диапазон колонок с A по B.
Примечание. Программа "LibreOffice Calc" не поддерживает эти свойства.
Смотрим пример:'''
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
# ==============================
# заставляем Excel печатать диапазон 
# строк со 2 по 3 на каждом листе
ws.print_title_rows = '2:3' 
# заставляем Excel печатать диапазон  
# колонок `A:A` на каждом листе
ws.print_title_cols = 'A:A'
# ==============================
# сохраняем и смотрим результат
wb.save('print.xlsx')



'''Определенные имена для доступа к диапазонам ячеек.'''
'''https://docs-python.ru/packages/modul-openpyxl/imenovannye-diapazony-jacheek/'''
'''Модуль openpyxl поддерживает определенные имена для более простого доступа к ячейке, диапазону ячеек или формулы.
Спецификация Excel говорит следующее об "определенных именах":
Определенные имена - это описательный текст, который используется для представления ячейки, диапазона ячеек, формулы или постоянного значения.
Это означает, что они очень слабо определены. Они могут содержать константу, формулу, ссылку на одну ячейку, ссылку на диапазон ячеек или несколько диапазонов ячеек на разных листах. Они определяются глобально для книги и доступны из атрибута Workbook.defined_names, который представляет собой словарный объект.
Создание именованных диапазонов.'''
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

wb = Workbook()
new_range = DefinedName('newrange', attr_text='Sheet!$A$1:$A$5')
wb.defined_names.append(new_range)

# локальный именованный диапазон 
# (действителен только для определенного листа)
sheetid = wb.sheetnames.index('Sheet')
private_range = DefinedName('privaterange', attr_text='Sheet!$A$6', localSheetId=sheetid)
wb.defined_names.append(private_range)
# этот локальный диапазон не может быть извлечен 
# из глобальных определенных имен
assert('privaterange' not in wb.defined_names)

# для получения локальных диапазонов 
# должна быть предоставлена область:
print(wb.defined_names.localnames(sheetid))
print(wb.defined_names.get('privaterange', sheetid).attr_text)

# код выведет
# ['privaterange']
# Sheet!$A$6

'''Пример доступа к диапазону ячеек.
Доступ к диапазону под названием my_range:'''
# получаем диапазон ячеек из словаря `defined_names`
my_range = wb.defined_names['my_range']
# если `my_range` содержит диапазон ячеек, то атрибут назначения не равен None

# следующее выражение возвращает генератор 
# кортежей (название листа, диапазон ячеек)
dests = my_range.destinations

cells = []
for title, coord in dests:
    ws = wb[title]
    cells.append(ws[coord])


'''Пример доступа к диапазону ячеек.
Доступ к диапазону под названием my_range:'''
# получаем диапазон ячеек из словаря `defined_names`
my_range = wb.defined_names['my_range']
# если `my_range` содержит диапазон ячеек, то атрибут назначения не равен None
# следующее выражение возвращает генератор кортежей (название листа, диапазон ячеек)
dests = my_range.destinations
cells = []
for title, coord in dests:
    ws = wb[title]
    cells.append(ws[coord])



'''Работа с объектом Table() модуля openpyxl в Python.
https://docs-python.ru/packages/modul-openpyxl/obekt-worksheet-table/'''

# --------------------------------------------------------------
'''Равнозначные записи'''
cells_range = ws['A3':f'A{max_row}']
for cel in cells_range:
    cel[0].border = bd


for col in range(1, 3 + 1):
    for row in range(3, max_row + 1):
        cel = ws.cell(row, col)
        cel.border = bd


for col in ws.iter_cols(min_row=3, max_col=max_column, max_row=max_row):
    for cell in col:
        cell.border = bd
# --------------------------------------------------------------

'''Вставка/удаление строк/столбцов, перемещение диапазона ячеек.'''

# Вставка строк и столбцов.
# Удаление строк и столбцов.
# Перемещение диапазона ячеек.
# Вставка строк и столбцов.

# Модуль openpyxl поддерживает вставку строк или столбцов. Что бы произвести указанные действия, необходимо использовать соответствующие методы экземпляра рабочего листа Worksheet:
# Worksheet.insert_cols(idx, amount=1): вставляет столбец или столбцы перед col==idx. Аргумент amount - количество добавляемых столбцов.
# Worksheet.insert_rows(idx, amount=1): вставляет строку или строки перед row==idx. Аргумент amount - количество добавляемых строк.
# По умолчанию вставляется одна строка или столбец. Например, чтобы вставить одну строку перед существующей 7-ой строкой необходимо вызвать ws.insert_rows(7).
# Пример:
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# создадим произвольные данные
data = [[row*col for col in range(1, 16)] for row in range(1, 31)]
# добавляем данные на активный лист
for row in data:
    ws.append(row)

# вставим 3 новые строки перед 
# существующей 7-ой строкой
ws.insert_rows(7, 3)
# сохраняем и смотрим
wb.save('test.xlsx')

'''# Удаление строк и столбцов.'''
# Что бы удалить строки или столбцы, используйте следующие методы экземпляра рабочего листа Worksheet:
# Worksheet.delete_cols(): удаляет столбец или столбцы, начиная с col==idx. Аргумент amount - количество удаляемых столбцов.
# Worksheet.delete_rows(): удаляет строку или строки, начиная с row==idx. Аргумент amount - количество удаляемых строк.
# По умолчанию удаляется одна строка или столбец. Например, чтобы удалить столбцы в диапазоне F:H необходимо вызвать ws.delete_cols(6, 3).
# Пример - продолжение предыдущего:
# удалим 3 столбцы в диапазоне `F:H`
ws.delete_cols(6, 3)
# сохраняемся и открываем файл
wb.save('test.xlsx')
# Примечание. При вставке или удалении строк или столбцов модуль openpyxl не управляет зависимостями, такими как формулы, таблицы, диаграммы и т.д. 
# Считается, что это выходит за рамки библиотеки, которая фокусируется на управлении форматом файла. 
# В общем, клиентский код должен реализовывать необходимую функциональность в любом конкретном случае.

'''# Перемещение диапазона ячеек.'''
# Модуль openpyxl обеспечивает перемещение диапазонов ячеек внутри листа методом:
# Worksheet.move_range(cell_range, rows=0, cols=0, translate=False).
# Этот метод перемещает диапазон ячеек cell_range на количество строк rows и/или столбцов cols:
# вниз, если rows > 0, и вверх, если rows < 0,
# вправо, если cols > 0, и влево, если cols < 0.
# Существующие ячейки будут перезаписаны. Формулы и ссылки обновляться не будут.
# Пример:
ws.move_range("D4:F10", rows=-1, cols=2)
# Это приведет к перемещению ячеек в диапазоне ячеек D4:F10 вверх на одну строку и вправо на два столбца. Ячейки будут перезаписаны всеми существующими ячейками.
# Если ячейки содержат формулы, то openpyxl может транслировать их, но, поскольку это не всегда то, что нужно, по этому этот функционал умолчанию отключен. 
# Кроме того, будут транслированы только формулы в самих ячейках. Ссылки на ячейки из других ячеек или определенные имена обновляться не будут. 
# Для этого можно использовать переводчик формул синтаксического анализа:
ws.move_range("G4:H10", rows=1, cols=1, translate=True)
# Это приведет к перемещению относительных ссылок в формулах в диапазоне на одну строку и один столбец.


from openpyxl import load_workbook

wb = load_workbook('width-height.xlsx')
ws = wb.active

# для строк
for i in range(1, ws.max_row+1):
    # если высота строки не изменялась программно
    # или вручную то `rh` будет присваиваться `None` 
    rh = ws.row_dimensions[i].height
    # по умолчанию высота строки равна 15 единицам
    row_heights = 15 if rh is None else rh
    print(f'Строка {i} имеет высоту {row_heights}')

# ну и для колонок
from openpyxl.utils.cell import get_column_letter
for i in range(1, ws.max_column+1):
    # преобразовываем индекс столбца в его букву
    letter = get_column_letter(i)
    # получаем ширину столбца
    col_width = ws.column_dimensions[letter].width


'''====================================================================='''
'''Копирование строк из одного листа в другой'''
# from copy import copy
# from openpyxl import load_workbook

# def copy_cell(src_sheet, src_row, src_col, 
#               tgt_sheet, tgt_row, tgt_col,
#               copy_style=True):
#     cell = src_sheet.cell(src_row, src_col)
#     new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
#     if cell.has_style and copy_style:
#         new_cell._style = copy(cell._style)

# foldrer = r"C:\vxvproj\tnnc-Excel\collectorExcel\collectorApp\temp"
# filename = r"\888888.xlsx"
# failstart = foldrer + filename
# wb = load_workbook(failstart)
# ws1 = wb['Лист1']
# ws2 = wb['Лист2']


# ws1_last_row = ws1.max_row

# for i, row in enumerate(ws1.iter_rows(min_row=2, max_row=3)):
# # for i, row in enumerate(ws2.iter_rows(min_row=2, max_col=3, max_row=1, values_only=True), 1):
#     for cell in row:
#         copy_cell(ws1, cell.row, cell.column, 
#                   ws2, ws2_last_row+i, cell.column)

# outfail = foldrer + r"\999.xlsx"
# wb.save(outfail)
# Excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
# wb = Excel.Workbooks.Open(outfail)

# Excel.Visible = 1
# Excel.DisplayAlerts = False
'''====================================================================='''
'''====================================================================='''
'''Копировоание строк из одного файла в другой'''
# from openpyxl import Workbook

# from copy import copy
# from openpyxl import load_workbook

# def copy_cell(src_sheet, src_row, src_col, 
#               tgt_sheet, tgt_row, tgt_col,
#               copy_style=True):
#     cell = src_sheet.cell(src_row, src_col)
#     new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
#     # if cell.has_style and copy_style:
#     #     new_cell._style = copy(cell._style)
#     new_cell.font = copy(cell.font)
#     new_cell.alignment = copy(cell.alignment)

# foldrer = r"C:\vxvproj\tnnc-Excel\collectorExcel\collectorApp\temp"
# filename = r"\888888.xlsx"
# failstart = foldrer + filename
# outfail = foldrer + r"\999.xlsx"
# print(outfail)

# wbcopy = load_workbook(failstart)
# wbcopy2 = load_workbook(outfail)
# wbxxx = Workbook()
# # wbxxx = load_workbook(foldrer + r"\pppppppppppppp.xlsx")

# ws1 = wbcopy['Лист1']
# ws2 = wbcopy2['Лист1']
# # ws3 = wbxxx.create_sheet("Лист1")
# ws3 = wbxxx.active

# for iii, u  in enumerate([ws1, ws2]):
#     ws3_last_row = ws3.max_row

#     for i, row in enumerate(u.iter_rows(min_row=1, max_row=ws1.max_row), start=1):
#         print("i = ", i)
#         ws2_last_row = ws2.max_row + 1
#         for cell in row:
#             copy_cell(u, cell.row, cell.column, 
#                     ws3, ws3_last_row + i, cell.column)

# wbxxx.save(foldrer + r"\pppppppppppppp.xlsx")
# wbcopy.close()
# wbcopy2.close()
# wbxxx.close()


# Excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
# wb = Excel.Workbooks.Open(foldrer + r"\pppppppppppppp.xlsx")

# Excel.Visible = 1
# Excel.DisplayAlerts = False
'''====================================================================='''

